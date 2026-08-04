"""Two evidence workbooks, one per optional chart.

Both share the same three-column shape -- item, frequency, and the interviews
behind it -- but they are built independently, because the two charts stand on
different kinds of evidence:

    Circumplex workbook    Every motivator/barrier's `frequency` is the
    (factor_circumplex)    human-coded count in config/motivators_barriers.
                           yaml -- AUTHORITATIVE, never recomputed here. The
                           config carries no per-person attribution, so the
                           names column is DERIVED by matching each factor's
                           cue lexicon against answer text. It is a looser
                           instrument than the coding and will not always
                           agree with the coded count.

    Opportunity workbook   Both the `mentions` count and the roster of names
    (opportunity_priority  come from the coded analysis, so the names column
     _matrix)              is the coded roster itself, not a re-detection.

Neither workbook reads the other's data: a change to the cue lexicons cannot
move an opportunity row, and a change to the opportunity roster cannot move a
motivator row.
"""

from __future__ import annotations

import logging
from pathlib import Path

from .models import RespondentProfile

logger = logging.getLogger(__name__)


def _person(respondent_id: str) -> str:
    """Viveka's workbook holds two response sheets, so her profile id is
    suffixed; the coding names each person once."""
    return respondent_id.split("__")[0]


def _display(name: str) -> str:
    return name.replace("_", " ").strip()


def _write(df, path: Path, sheet_name: str, widths: list[int]) -> None:
    """One sheet, frozen header, wrapped name column. The names column runs to
    a couple of hundred characters, so it is wrapped rather than left to spill
    across the sheet."""
    import pandas as pd
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        header_fill = PatternFill("solid", fgColor="EFD9FF")
        for col, width in enumerate(widths, start=1):
            letter = get_column_letter(col)
            ws.column_dimensions[letter].width = width
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.freeze_panes = "A2"


# --------------------------------------------------------------------------
# Workbook 1 -- circumplex (motivators & barriers)
# --------------------------------------------------------------------------

def circumplex_report_dataframe(profiles: list[RespondentProfile], config):
    """One row per motivator/barrier: the factor, its coded frequency, and the
    interviews its cues fire in."""
    import pandas as pd

    from .drivers import collect_driver_hits, load_driver_items

    items = load_driver_items(config)
    hits = collect_driver_hits(profiles, config)

    rows = []
    for item in items:
        people = sorted({_display(_person(r["respondent_id"]))
                         for r in hits.get(item.key, [])})
        rows.append({
            "Motivator / Barrier": f"{item.code} — {item.label}",
            "Frequency": item.frequency,
            "Interviews (derived from transcript cue matching)":
                ", ".join(people) if people else "no cue match in transcripts",
        })
    return pd.DataFrame(rows)


def write_circumplex_report(profiles: list[RespondentProfile], config, path: Path) -> bool:
    """Writes the motivator/barrier evidence workbook behind
    factor_circumplex.png. Returns False if nothing is configured."""
    from .drivers import load_driver_items

    if not load_driver_items(config):
        logger.warning("No motivators/barriers configured; skipping circumplex report.")
        return False
    df = circumplex_report_dataframe(profiles, config)
    _write(df, path, "Motivators & Barriers", [58, 12, 110])
    return True


# --------------------------------------------------------------------------
# Workbook 2 -- opportunity priority matrix
# --------------------------------------------------------------------------

def opportunity_report_dataframe(profiles: list[RespondentProfile], config):
    """One row per opportunity: the theme, its coded mention count, and the
    interviews the coding names."""
    import pandas as pd

    from .opportunities import _norm, load_opportunities

    known = {_norm(_person(p.respondent_id)) for p in profiles}

    rows = []
    for opp in load_opportunities(config):
        # A coded name with no matching workbook is flagged inline rather than
        # dropped -- it means the coding and the file set disagree, which is
        # exactly the thing this sheet exists to surface.
        names = [f"{n} (no interview file)" if _norm(n) not in known else n
                 for n in opp.mentioned_by]
        rows.append({
            "Opportunity": f"{opp.code} — {opp.title}",
            "Frequency": opp.mentions,
            "Interviews (from the coded roster)":
                ", ".join(names) if names else "no names recorded",
        })
    return pd.DataFrame(rows)


def write_opportunity_report(profiles: list[RespondentProfile], config, path: Path) -> bool:
    """Writes the opportunity evidence workbook behind
    opportunity_priority_matrix.png. Returns False if nothing is configured."""
    from .opportunities import load_opportunities

    if not load_opportunities(config):
        logger.warning("No opportunities configured; skipping opportunity report.")
        return False
    df = opportunity_report_dataframe(profiles, config)
    _write(df, path, "Opportunities", [58, 12, 110])
    return True
