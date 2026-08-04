"""Stage 16/17/20 (spec sections 14, 15, 16, 40, 41): CSV/Excel exports."""

from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .config_loader import (
    THEMES,
    group_for_persona,
    persona_group_display_names,
    persona_groups,
)
from .models import RespondentProfile

logger = logging.getLogger(__name__)

THEME_ORDER = THEMES

# Header/footer-label color, sampled from the reference "Interview Response
# Evaluation" table this export is modeled after.
HEADER_PURPLE = "7500C0"

# A red -> amber -> green scale for 1-5 theme/average scores, shared by the
# Excel export and the PNG heatmap so the two outputs always agree on what a
# given score "feels like". Each tier has a saturated "face" color (the smiley
# itself) and a pastel "bg" color (the cell background) -- the same two-tone
# look as the reference "Interview Response Evaluation" table.
#
# Six tiers, ordered highest -> lowest.
#
# The band count is a balance between two opposite failures, both of which this
# scale has actually hit:
#
#   Too many bands and neighbours become indistinguishable. An earlier
#   eight-tier version graded hue and mouth curvature in small steps and
#   produced pairs nobody could rank -- two yellows differing only in how far
#   the mouth turned up, and three greens likewise.
#
#   Too few and the bands stop discriminating. A five-tier version put 2.5-3.5
#   in one "Neutral" band, which swallowed roughly two thirds of a real matrix
#   because scored themes cluster hard around the midpoint.
#
# Six works because the bands are cut where the data actually falls (a narrower
# 2.6-3.2 middle, splitting the old neutral block) while every tier still owns
# a hue from a DIFFERENT family -- red / orange / yellow / olive / green / dark
# green. Only one band is yellow. Expression is a second, independent cue:
# deep frown / frown / flat bar / slight smile / smile / open grin, with sad
# brows on the worst band and arced eyes on the best.
#
# Caution when reading the two middle bands: theme scores in this pipeline
# carry low evidence coverage, so a 0.4-wide band is near the limit of what the
# underlying numbers support. Treat a one-band difference as suggestive and
# check `evidence_status` before acting on it.
#
# `label` is the human-readable band name shown in the heatmap's scale legend,
# and `direction` states the reading outright so nothing rests on a reader
# interpreting a face correctly.
SMILEY_TIERS = [
    {"min": 4.0, "face": "2E7D32", "bg": "A5D6A7", "emoji": "\U0001F604",
     "curve": 1.0, "label": "Strong", "direction": "better"},
    {"min": 3.6, "face": "7CB342", "bg": "DCEDC8", "emoji": "\U0001F60A",
     "curve": 0.7, "label": "Good", "direction": "better"},
    {"min": 3.2, "face": "C0CA33", "bg": "F0F4C3", "emoji": "\U0001F642",
     "curve": 0.35, "label": "Fair", "direction": "better"},
    {"min": 2.6, "face": "FBC02D", "bg": "FFF9C4", "emoji": "\U0001F610",
     "curve": 0.0, "label": "Neutral", "direction": "middle"},
    {"min": 2.0, "face": "EF6C00", "bg": "FFCC80", "emoji": "\U0001F641",
     "curve": -0.55, "label": "Weak", "direction": "worse"},
    {"min": float("-inf"), "face": "C62828", "bg": "FFCDD2", "emoji": "\U0001F61E",
     "curve": -1.0, "label": "Poor", "direction": "worse"},
]


def _smiley_tier(score) -> dict | None:
    """Look up the display tier for a 1-5 score.

    Returns None for missing/NaN scores (no respondents in that cell). Also
    guards against openpyxl round-tripping a blank/NaN cell as ''."""
    if score is None or score == "":
        return None
    try:
        score = float(score)
    except (TypeError, ValueError):
        return None
    if pd.isna(score):
        return None
    for tier in SMILEY_TIERS:
        if score >= tier["min"]:
            return tier
    return SMILEY_TIERS[-1]


def _draw_score_face(ax, cx, cy, tier, size=0.30):
    """Draws one tier's face onto `ax`.

    Module level, not nested inside the heatmap, so the heatmap grid and the
    standalone score-scale legend render identical faces by construction --
    the legend is only useful if the faces in it are the same ones in the
    chart.

    Every cue is redundant with every other -- hue, mouth shape, mouth
    weight and brow treatment all move together -- so the ranking survives
    greyscale printing, small thumbnails, and colour-blind readers. The
    previous version graded only hue and curvature in small steps, which
    left neighbouring tiers indistinguishable.
    """
    from matplotlib.patches import Arc, Circle, Wedge

    face_color = f"#{tier['face']}"
    curve = tier["curve"]
    ax.add_patch(Circle((cx, cy), size, facecolor=face_color, edgecolor="none", zorder=3))

    eye_dx, eye_dy, eye_r = size * 0.36, size * 0.28, size * 0.10
    if curve >= 0.9:
        # Top tier: happy arced eyes, unmistakable even in miniature.
        for sx in (-1, 1):
            ax.add_patch(Arc((cx + sx * eye_dx, cy + eye_dy), size * 0.40, size * 0.34,
                              theta1=20, theta2=160, color="black", linewidth=1.8, zorder=4))
    else:
        for sx in (-1, 1):
            ax.add_patch(Circle((cx + sx * eye_dx, cy + eye_dy), eye_r, facecolor="black", zorder=4))
        if curve <= -0.9:
            # Bottom tier only: sad brows, sloping down toward the outside
            # of the face. A third cue on top of hue and mouth, so the
            # worst band is identifiable even in greyscale.
            for sx in (-1, 1):
                ax.plot([cx + sx * size * 0.56, cx + sx * size * 0.20],
                        [cy + size * 0.44, cy + size * 0.62],
                        color="black", linewidth=1.8, zorder=4, solid_capstyle="round")

    mouth_w = size * 1.0
    if abs(curve) < 0.05:
        # Neutral: a flat bar, visually distinct from any arc.
        ax.plot([cx - mouth_w / 2, cx + mouth_w / 2], [cy - size * 0.34] * 2,
                 color="black", linewidth=2.2, zorder=4, solid_capstyle="round")
    elif curve >= 0.9:
        # Top tier: a filled open grin rather than a line, so it reads as
        # categorically different from the merely-smiling tier below it.
        ax.add_patch(Wedge((cx, cy - size * 0.04), size * 0.62, 200, 340,
                             facecolor="black", edgecolor="none", zorder=4))
    else:
        # Arcs are positioned by their EXTREME point (deepest point of a
        # smile, highest point of a frown) rather than by their centre, so
        # a deeper curve grows downward/upward from a fixed lip line
        # instead of scaling about the middle of the face. Sizing by centre
        # let the -1.0 frown grow tall enough to swallow the eyes and read
        # as a smile.
        mouth_h = max(size * 0.28, size * 0.62 * abs(curve))
        if curve > 0:
            center_y = cy - size * 0.46 + mouth_h / 2
            theta1, theta2 = 200, 340
        else:
            center_y = cy - size * 0.24 - mouth_h / 2
            theta1, theta2 = 20, 160
        ax.add_patch(Arc((cx, center_y), mouth_w, mouth_h,
                          theta1=theta1, theta2=theta2,
                          color="black", linewidth=2.2, zorder=4))


def _safe_mean(values) -> float | None:
    vals = [v for v in values if v is not None and not (isinstance(v, float) and pd.isna(v))]
    return round(sum(vals) / len(vals), 2) if vals else None

ROSTER_COLUMNS = [
    "Brand Persona", "respondent_id", "source_file",
]

RESPONDENT_COLUMNS = [
    "respondent_id", "respondent_id_source", "respondent_id_confidence",
    "source_file", "sheet_name", "inferred_role", "original_persona_sheet_label",
    # The persona as reported (one of the two merged personas) alongside the
    # persona the classifier actually decided on. Every other column on this
    # row -- the margin, the reasoning, the gate flags -- describes the
    # four-way decision, so dropping it would leave them unexplainable.
    "reported_brand_persona", "brand_persona",
    "classification_confidence", "engagement_gate_satisfied",
    "confidence", "evidence_coverage", "evidence_strength",
    "persona_margin", "answer_completeness", "classification_consistency",

    # Engagement / relationship state -- the facts the persona gate is decided
    # on, exported so a reviewer can check the classification without rerunning
    # anything.
    "sections_answered", "has_jabc_experience", "has_current_jabc_engagement",
    "has_lapsed_jabc_engagement", "intends_to_continue_jabc", "is_aware_of_jabc",
    "runs_external_programs", "explores_external_programs",
    "has_positive_external_experience", "has_negative_external_experience",
    "is_hesitant_about_external_programs", "has_no_external_experience",
    "has_current_barriers",
    "jabc_recency_signal", "external_disposition_signal", "engagement_evidence",

    "awareness_score", "awareness_evidence_strength", "awareness_evidence_status",
    "understanding_score", "understanding_evidence_strength", "understanding_evidence_status",
    "trust_score", "trust_evidence_strength", "trust_evidence_status",
    "relevance_score", "relevance_evidence_strength", "relevance_evidence_status",
    "ease_of_engagement_score", "ease_of_engagement_evidence_strength", "ease_of_engagement_evidence_status",
    "advocacy_score", "advocacy_evidence_strength", "advocacy_evidence_status",

    "average_score", "overall_sentiment_label", "overall_sentiment_score",
    "manual_review_flag", "manual_review_reason", "classification_reasoning",
]

EVIDENCE_COLUMNS = [
    "respondent_id", "source_file", "sheet_name", "question_number", "question_category",
    "question_text", "answer_excerpt", "theme", "evidence_type", "evidence_direction",
    "evidence_strength", "question_context_weight", "matched_phrases", "matched_keywords",
    "negation_detected", "intensifier_detected", "sentiment_score",
]

VALIDATION_COLUMNS = [
    "respondent_id", "original_persona_sheet_label", "predicted_brand_persona",
    "classification_match", "confidence", "classification_confidence",
    "engagement_gate_satisfied", "manual_review_flag",
]

# Which of the two reported Brand Personas each original sheet label belongs
# to. The labels were written under the older four-way naming, so the mapping
# is by substring on the label; anything unrecognized yields a null match
# rather than a false one.
_LABEL_GROUP_HINTS = {
    "established_supporters": ["active", "lapsed", "familiar-history"],
    "potential_adopters": ["familiar-nohistory", "famunfam", "unfamiliar"],
}


def _mean_theme_evidence(profile: RespondentProfile, field: str) -> float:
    vals = [getattr(ts, field) for ts in profile.theme_scores.values()]
    return round(sum(vals) / len(vals), 3) if vals else 0.0


def _reported_persona_name(profile: RespondentProfile, config=None) -> str:
    """The merged persona a respondent is reported under. Falls back to the
    classification persona's own name if it belongs to no group, so a
    misconfigured grouping shows up in the output rather than as a blank."""
    group_key = group_for_persona(profile.brand_persona, config)
    if group_key is None:
        return profile.persona_display_name
    return persona_group_display_names(config).get(group_key, group_key)


def respondent_scores_dataframe(profiles: list[RespondentProfile], config=None) -> pd.DataFrame:
    rows = []
    for p in profiles:
        row = {
            "respondent_id": p.respondent_id,
            "respondent_id_source": p.respondent_id_source,
            "respondent_id_confidence": p.respondent_id_confidence,
            "source_file": p.source_file,
            "sheet_name": p.sheet_name,
            "inferred_role": p.inferred_role,
            "original_persona_sheet_label": p.original_persona_sheet_label,
            "reported_brand_persona": _reported_persona_name(p, config),
            "brand_persona": p.persona_display_name,
            "classification_confidence": p.classification_confidence,
            "engagement_gate_satisfied": p.engagement_gate_satisfied,
            "confidence": p.confidence,
            "evidence_coverage": _mean_theme_evidence(p, "evidence_coverage"),
            "evidence_strength": _mean_theme_evidence(p, "evidence_strength"),
            "persona_margin": p.persona_margin,
            "answer_completeness": p.answer_completeness,
            "classification_consistency": p.classification_consistency,
        }
        row["sections_answered"] = "".join(p.sections_answered)
        for flag in ["has_jabc_experience", "has_current_jabc_engagement",
                     "has_lapsed_jabc_engagement", "intends_to_continue_jabc",
                     "is_aware_of_jabc", "runs_external_programs",
                     "explores_external_programs", "has_positive_external_experience",
                     "has_negative_external_experience",
                     "is_hesitant_about_external_programs", "has_no_external_experience",
                     "has_current_barriers"]:
            row[flag] = p.behavioral_state.get(flag, False)
        row["jabc_recency_signal"] = p.engagement_signals.get("jabc_recency", 0.0)
        row["external_disposition_signal"] = p.engagement_signals.get("external_disposition", 0.0)
        # Flat citation list so each gate decision can be traced to the answers
        # that drove it without opening the evidence workbook.
        row["engagement_evidence"] = " | ".join(
            f"{flag}: {excerpt}"
            for flag, excerpts in sorted(p.engagement_evidence.items())
            for excerpt in excerpts[:2]
        )[:1500]

        for theme in THEME_ORDER:
            ts = p.theme_scores.get(theme)
            row[f"{theme}_score"] = ts.final_score if ts else None
            row[f"{theme}_evidence_strength"] = ts.evidence_strength if ts else None
            row[f"{theme}_evidence_status"] = ts.evidence_status if ts else None

        row["average_score"] = round(p.average_score(), 3) if p.average_score() is not None else None
        row["overall_sentiment_label"] = p.overall_sentiment_label
        row["overall_sentiment_score"] = p.overall_sentiment_score
        row["manual_review_flag"] = p.manual_review_flag
        row["manual_review_reason"] = p.manual_review_reason
        row["classification_reasoning"] = p.classification_reasoning
        rows.append(row)

    return pd.DataFrame(rows, columns=RESPONDENT_COLUMNS)


def evidence_dataframe(profiles: list[RespondentProfile]) -> pd.DataFrame:
    rows = []
    for p in profiles:
        for rec in p.evidence_records:
            if rec.theme == "__behavioral__":
                continue  # behavioral flags are reported in the respondent table, not here
            rows.append({
                "respondent_id": rec.respondent_id,
                "source_file": rec.source_file,
                "sheet_name": rec.sheet_name,
                "question_number": rec.question_number,
                "question_category": rec.question_category,
                "question_text": rec.question_text,
                "answer_excerpt": rec.evidence_text,
                "theme": rec.theme,
                "evidence_type": rec.evidence_type,
                "evidence_direction": rec.direction,
                "evidence_strength": rec.strength,
                "question_context_weight": rec.question_context_weight,
                "matched_phrases": ", ".join(rec.matched_phrases),
                "matched_keywords": ", ".join(rec.matched_keywords),
                "negation_detected": rec.negation_detected,
                "intensifier_detected": rec.intensifier_detected,
                "sentiment_score": round(rec.sentiment_support, 3),
            })
    return pd.DataFrame(rows, columns=EVIDENCE_COLUMNS)


def validation_dataframe(profiles: list[RespondentProfile], config=None) -> pd.DataFrame:
    rows = []
    for p in profiles:
        original = p.original_persona_sheet_label
        # Reported at the two-persona level, the same unit as every other
        # output. The four-way gate result is an internal step of the
        # classification and is not a reportable persona.
        predicted = _reported_persona_name(p, config)
        # A loose textual match: does the predicted persona's key phrase
        # appear in the original sheet label? This is diagnostic only, per
        # spec section 33 -- it never feeds back into classification.
        match = _loose_label_match(original, p.brand_persona, config)
        rows.append({
            "respondent_id": p.respondent_id,
            "original_persona_sheet_label": original,
            "predicted_brand_persona": predicted,
            "classification_match": match,
            "confidence": p.confidence,
            "classification_confidence": p.classification_confidence,
            "engagement_gate_satisfied": p.engagement_gate_satisfied,
            "manual_review_flag": p.manual_review_flag,
        })
    return pd.DataFrame(rows, columns=VALIDATION_COLUMNS)


def persona_roster_dataframe(profiles: list[RespondentProfile], config=None) -> pd.DataFrame:
    """One row per respondent, grouped under its Brand Persona -- the roster of
    "who is in which persona" (companion view to the persona-level averages in
    the Brand Health Matrix). Respondents are listed alphabetically within a
    persona; the gate the classifier passed through to get there is an internal
    step and is not reported here."""
    rows = []
    display = persona_group_display_names(config)
    for group_key, members in persona_groups(config).items():
        member_profiles = sorted(
            (p for p in profiles if p.brand_persona in members),
            key=lambda p: p.respondent_id,
        )
        for p in member_profiles:
            rows.append({
                "Brand Persona": display.get(group_key, group_key),
                "respondent_id": p.respondent_id,
                "source_file": p.source_file,
            })
    return pd.DataFrame(rows, columns=ROSTER_COLUMNS)


def _loose_label_match(original_label: str, persona_key: str, config=None) -> bool | None:
    """Does the original sheet label point at the same reported persona the
    classifier chose? Matched at the two-persona level, because that is the
    unit the validation file reports. A label that names no group -- or names
    both -- is not evidence either way, so it returns None instead of False."""
    label = (original_label or "").lower()
    hit = [
        group for group, hints in _LABEL_GROUP_HINTS.items()
        if any(h in label for h in hints)
    ]
    if len(hit) != 1:
        return None
    predicted_group = group_for_persona(persona_key, config)
    if predicted_group is None:
        return None
    return hit[0] == predicted_group


def write_csv(df: pd.DataFrame, path: Path) -> None:
    df.to_csv(path, index=False)


def write_matrix_excel(df: pd.DataFrame, path: Path, roster_df: pd.DataFrame | None = None,
                         config=None) -> None:
    """Writes the Brand Health Matrix modeled on the "Interview Response
    Evaluation" reference table: a purple header/footer-label band, each
    theme score rendered as a smiley face on a pastel tier-colored cell
    (color + expression graded across SMILEY_TIERS) instead of a bare
    number. Two summary views are appended so nothing is lost by dropping
    the raw numbers from the main grid:
      - an "Average" row at the bottom: the mean of each theme column
        across personas, as plain numbers on the same pastel tier colors.
      - an "Average Score" column at the end: the mean across themes for
        that persona, as a plain italic number (row-wise average).

    Cells listed in matrix.suppressed_cells (ease of engagement and advocacy
    for Potential Adopters) are left genuinely empty on a white
    fill -- no smiley, no dash. They are deliberately not applicable rather
    than merely missing, and an en dash on a tinted cell reads as "we tried and
    got nothing", which is a different and misleading claim.

    If `roster_df` is given, it is written as a second sheet ("Persona
    Members") showing which respondents fall into each Brand Persona."""
    from .matrix import suppressed_cells

    blanked = suppressed_cells(config)
    persona_names = df["Brand Persona"].tolist()
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    theme_cols = ["Awareness", "Understanding", "Trust", "Relevance",
                  "Ease of Engagement", "Advocacy"]
    col_order = ["Brand Persona"] + theme_cols + ["Respondent Count", "Average Score"]
    disp_df = df[col_order].copy()

    summary_row = {"Brand Persona": "Average"}
    for col in theme_cols:
        # Column average over the personas the theme actually applies to. A
        # suppressed cell is already None in the matrix, and _safe_mean skips
        # None, so this stays consistent with the row averages by construction.
        summary_row[col] = _safe_mean(disp_df[col])
    summary_row["Respondent Count"] = int(disp_df["Respondent Count"].sum())
    summary_row["Average Score"] = _safe_mean(summary_row[col] for col in theme_cols)
    disp_df = pd.concat([disp_df, pd.DataFrame([summary_row])], ignore_index=True)

    n_persona_rows = len(df)
    summary_row_idx = len(disp_df)  # 1-indexed, header is row 1

    header_fill = PatternFill(start_color=HEADER_PURPLE, end_color=HEADER_PURPLE, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_font_italic = Font(bold=True, italic=True, color="FFFFFF")
    thin_border = Border(*(Side(style="thin", color="BBBBBB"),) * 4)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        disp_df.to_excel(writer, sheet_name="Brand Health Matrix", index=False)
        ws = writer.sheets["Brand Health Matrix"]

        for col_idx, col_name in enumerate(disp_df.columns, start=1):
            header_cell = ws.cell(row=1, column=col_idx)
            header_cell.fill = header_fill
            header_cell.font = header_font_italic if col_name == "Average Score" else header_font
            header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            max_len = max([len(str(col_name))] + [len(str(v)) for v in disp_df[col_name]])
            ws.column_dimensions[get_column_letter(col_idx)].width = min(28, max_len + 4)

        avg_score_col_idx = col_order.index("Average Score") + 1

        for row_offset in range(n_persona_rows):
            excel_row = row_offset + 2  # +1 for header, +1 for 1-indexing
            persona_name = persona_names[row_offset] if row_offset < len(persona_names) else ""
            for col_name in theme_cols:
                col_idx = col_order.index(col_name) + 1
                cell = ws.cell(row=excel_row, column=col_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                if (persona_name, col_name) in blanked:
                    cell.value = None
                    cell.fill = white_fill
                    cell.font = Font(size=14)
                    continue
                tier = _smiley_tier(cell.value)
                cell.value = tier["emoji"] if tier else "–"
                cell.font = Font(size=14)
                if tier:
                    cell.fill = PatternFill(start_color=tier["bg"], end_color=tier["bg"], fill_type="solid")

            avg_cell = ws.cell(row=excel_row, column=avg_score_col_idx)
            tier = _smiley_tier(avg_cell.value)
            if tier:
                avg_cell.fill = PatternFill(start_color=tier["bg"], end_color=tier["bg"], fill_type="solid")
            avg_cell.font = Font(bold=True, italic=True)
            avg_cell.border = thin_border
            ws.cell(row=excel_row, column=1).border = thin_border
            ws.cell(row=excel_row, column=col_order.index("Respondent Count") + 1).border = thin_border

        # "Average" summary row: the label cell matches the purple header
        # band (as in the reference table); each theme/average cell keeps
        # its own pastel tier color but with plain bold numbers rather than
        # a smiley, so it reads as a totals row rather than another score.
        excel_summary_row = summary_row_idx + 1  # +1 for header row
        label_cell = ws.cell(row=excel_summary_row, column=1)
        label_cell.fill = header_fill
        label_cell.font = header_font_italic
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        label_cell.border = thin_border

        grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        for col_name in theme_cols + ["Average Score"]:
            col_idx = col_order.index(col_name) + 1
            cell = ws.cell(row=excel_summary_row, column=col_idx)
            tier = _smiley_tier(cell.value)
            cell.font = Font(bold=True, italic=(col_name == "Average Score"))
            cell.border = thin_border
            cell.fill = PatternFill(start_color=tier["bg"], end_color=tier["bg"], fill_type="solid") if tier else grey_fill

        count_cell = ws.cell(row=excel_summary_row, column=col_order.index("Respondent Count") + 1)
        count_cell.font = Font(bold=True)
        count_cell.fill = grey_fill
        count_cell.border = thin_border

        if roster_df is not None:
            roster_df.to_excel(writer, sheet_name="Persona Members", index=False)
            roster_ws = writer.sheets["Persona Members"]
            for col_idx, col_name in enumerate(roster_df.columns, start=1):
                roster_ws.cell(row=1, column=col_idx).font = Font(bold=True)
                max_len = max([len(str(col_name))] + [len(str(v)) for v in roster_df[col_name]])
                roster_ws.column_dimensions[get_column_letter(col_idx)].width = min(40, max_len + 2)


def write_dataframe_excel(df: pd.DataFrame, path: Path, sheet_name: str = "Sheet1") -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.cell(row=1, column=col_idx).font = Font(bold=True)
            if len(df) > 0:
                max_len = max([len(str(col_name))] + [len(str(v)[:60]) for v in df[col_name]])
            else:
                max_len = len(str(col_name))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(40, max_len + 2)


def write_score_scale(path: Path) -> bool:
    """Writes the score-scale key as its own PNG.

    Kept out of the matrix image so the matrix stays a clean grid that can be
    dropped into a slide at any size. The key is a fixed reference -- it never
    changes with the data -- so it belongs in a separate asset that gets placed
    once, rather than re-rendered under every chart.

    Faces come from `_draw_score_face`, the same function the heatmap cells
    use, so the key cannot drift out of step with what it is explaining.
    """
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from matplotlib.patches import Rectangle
    except ImportError:
        logger.warning("matplotlib not available; skipping score scale legend.")
        return False

    n = len(SMILEY_TIERS)
    swatch_w, swatch_h = 2.6, 1.15

    fig, ax = plt.subplots(figsize=(1.75 * n + 0.6, 2.5))
    ax.set_xlim(-0.15, n * swatch_w + 0.15)
    ax.set_ylim(-0.45, swatch_h + 1.15)
    ax.set_aspect("equal")
    ax.axis("off")

    # Direction rail above the bands, so "which way is good" is stated rather
    # than left to the reader to infer from colour convention.
    rail_y = swatch_h + 0.42
    ax.text(0.0, rail_y, "worse", ha="left", va="center",
             fontsize=10, fontstyle="italic", color="#999999")
    ax.text(n * swatch_w, rail_y, "better", ha="right", va="center",
             fontsize=10, fontstyle="italic", color="#999999")
    ax.annotate("", xy=(n * swatch_w - 0.95, rail_y), xytext=(0.95, rail_y),
                 arrowprops=dict(arrowstyle="-|>", color="#999999", linewidth=1.2))

    # Rendered worst -> best, left to right, matching the rail above.
    for k, tier in enumerate(reversed(SMILEY_TIERS)):
        x0 = k * swatch_w
        ax.add_patch(Rectangle((x0, 0), swatch_w - 0.18, swatch_h,
                                facecolor=f"#{tier['bg']}", edgecolor="#BBBBBB",
                                linewidth=0.9, zorder=1))
        _draw_score_face(ax, x0 + 0.46, swatch_h / 2, tier, size=0.32)

        lo = tier["min"]
        higher = SMILEY_TIERS[n - 1 - k - 1]["min"] if k < n - 1 else None
        if lo == float("-inf"):
            band = f"below {higher:.1f}" if higher is not None else "lowest"
        elif higher is None:
            band = f"{lo:.1f} and up"
        else:
            band = f"{lo:.1f} – {higher:.1f}"

        ax.text(x0 + 0.92, swatch_h / 2 + 0.19, tier["label"], ha="left", va="center",
                 fontsize=11.5, fontweight="bold", color="#333333", zorder=2)
        ax.text(x0 + 0.92, swatch_h / 2 - 0.22, band, ha="left", va="center",
                 fontsize=10, color="#666666", zorder=2)

    ax.set_title("Brand Health Score Scale", fontsize=13, fontweight="bold",
                  color="#1a1a1a", pad=14)
    fig.savefig(path, dpi=150, bbox_inches="tight", pad_inches=0.25)
    plt.close(fig)
    return True


def write_heatmap(df: pd.DataFrame, path: Path, config=None) -> bool:
    """Optional heatmap (spec section 16), modeled on the "Interview
    Response Evaluation" reference table: a purple header band, each
    persona/theme cell filled with its pastel tier color and holding a
    procedurally-drawn flat smiley face (not an emoji character -- those
    don't render reliably through matplotlib's Agg backend) whose color and
    mouth curvature reflect SMILEY_TIERS. A purple-labeled "Average"
    summary row/column (plain numbers on the same pastel tier colors) is
    appended, mirroring the Excel export.

    Cells listed in matrix.suppressed_cells (ease of engagement and advocacy
    for Potential Adopters) are drawn as plain white -- no face,
    no dash. Those cells are not applicable rather than missing, and a dash on
    a tinted cell would read as a failed measurement instead of a question that
    was never put to that persona."""
    from .matrix import suppressed_cells
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from matplotlib.patches import Arc, Circle, Rectangle, Wedge
        import textwrap
    except ImportError:
        logger.warning("matplotlib not available; skipping optional heatmap.")
        return False

    theme_cols = ["Awareness", "Understanding", "Trust", "Relevance", "Ease of Engagement", "Advocacy"]
    personas = df["Brand Persona"].tolist()
    purple = f"#{HEADER_PURPLE}"
    blanked = suppressed_cells(config)

    def draw_face_cell(ax, col_idx, y, score, suppressed=False):
        cx, cy = col_idx + 0.5, y + 0.5
        if suppressed:
            ax.add_patch(Rectangle((col_idx, y), 1, 1, facecolor="white",
                                    edgecolor="#BBBBBB", linewidth=0.8, zorder=1))
            return
        tier = _smiley_tier(score)
        bg = f"#{tier['bg']}" if tier else "white"
        ax.add_patch(Rectangle((col_idx, y), 1, 1, facecolor=bg, edgecolor="#BBBBBB", linewidth=0.8, zorder=1))
        if tier is None:
            ax.text(cx, cy, "–", ha="center", va="center", fontsize=13, color="#999999", zorder=2)
            return
        _draw_score_face(ax, cx, cy, tier, size=0.32)

    def draw_summary_cell(ax, col_idx, y, value):
        tier = _smiley_tier(value)
        bg = f"#{tier['bg']}" if tier else "#EDEDED"
        ax.add_patch(Rectangle((col_idx, y), 1, 1, facecolor=bg, edgecolor="#BBBBBB", linewidth=0.8, zorder=1))
        has_value = value is not None and not (isinstance(value, float) and pd.isna(value))
        label = f"{value:.2f}" if has_value else "–"
        ax.text(col_idx + 0.5, y + 0.5, label, ha="center", va="center", fontsize=12,
                 fontweight="bold", fontstyle="italic", color="#333333", zorder=2)

    n_cols = len(theme_cols) + 1   # + Average column
    n_rows = len(personas) + 1     # + Average row
    header_h = 0.9

    fig, ax = plt.subplots(figsize=(1.7 * n_cols + 1.6, 1.1 * n_rows + 1.8))
    ax.set_xlim(-2.6, n_cols)
    ax.set_ylim(-0.6, n_rows + header_h + 0.3)
    ax.axis("off")

    def row_y(row_idx):
        """Row 0 (first persona) drawn at the top; summary row at the bottom."""
        return n_rows - 1 - row_idx

    # Purple header band, matching the reference table's header row.
    ax.add_patch(Rectangle((-2.6, n_rows), n_cols + 2.6, header_h, facecolor=purple, edgecolor="none", zorder=1))
    ax.text(-1.3, n_rows + header_h / 2, "Brand\nPersona", ha="center", va="center",
             fontsize=10, fontweight="bold", color="white", zorder=2)
    for j, col in enumerate(theme_cols):
        wrapped = "\n".join(textwrap.wrap(col, 10, break_long_words=False))
        ax.text(j + 0.5, n_rows + header_h / 2, wrapped, ha="center", va="center",
                 fontsize=9.5, fontweight="bold", color="white", zorder=2)
    ax.text(len(theme_cols) + 0.5, n_rows + header_h / 2, "Average\nof Personas", ha="center", va="center",
             fontsize=9.5, fontweight="bold", fontstyle="italic", color="white", zorder=2)

    for i, persona in enumerate(personas):
        y = row_y(i)
        ax.text(-1.3, y + 0.5, persona, ha="center", va="center", fontsize=10, fontweight="bold")
        for j, col in enumerate(theme_cols):
            draw_face_cell(ax, j, y, df.iloc[i][col], suppressed=(persona, col) in blanked)
        draw_summary_cell(ax, len(theme_cols), y, df.iloc[i]["Average Score"])

    summary_y = row_y(len(personas))
    ax.add_patch(Rectangle((-2.6, summary_y), 2.6, 1, facecolor=purple, edgecolor="#BBBBBB", linewidth=0.8, zorder=1))
    ax.text(-1.3, summary_y + 0.5, "Average\nof Themes", ha="center", va="center", fontsize=10,
             fontweight="bold", fontstyle="italic", color="white", zorder=2)
    theme_averages = [_safe_mean(df[col]) for col in theme_cols]
    for j, avg in enumerate(theme_averages):
        draw_summary_cell(ax, j, summary_y, avg)
    draw_summary_cell(ax, len(theme_cols), summary_y, _safe_mean(theme_averages))

    ax.set_title("JABC Brand Health Matrix", fontsize=15, fontweight="bold", pad=24)
    fig.savefig(path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    return True
