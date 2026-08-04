"""Generates synthetic JABC interview workbooks for testing and demoing the
pipeline. Mirrors the general shape described in the build spec: a
reference-only 'Question Bank' sheet plus one completed persona-specific
response sheet per workbook.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

HEADERS = ["Question Number", "Question Category", "Question", "Answer (record answers here)", "Notes"]

INTERVIEWS = {
    "Marci_Interview.xlsx": {
        "sheet": "P1 - Teacher-Active",
        "rows": [
            ("C1", "Familiarity & Perception", "How familiar are you with JABC? (1-5)", "5", ""),
            ("C2", "Familiarity & Perception", "What is the first thing that comes to mind when you hear JABC?",
             "JABC is excellent and the outreach team is really helpful and responsive.", ""),
            ("C4", "Familiarity & Perception", "Describe JABC to a colleague.",
             "It's hands-on financial literacy and entrepreneurship programming that aligns with our curriculum.", ""),
            ("D2", "Decision-Making", "What makes a program worth classroom time?",
             "Programs that connect directly to student needs and instructional goals, like JABC does.", ""),
            ("E1", "JABC Experience", "Describe your recent JABC experience.",
             "We had a wonderful, reliable experience this year. Registration was easy and communication was clear.", ""),
            ("I3", "Communication & Future Opportunities", "Would you recommend or champion JABC to other educators?",
             "Absolutely, I would 100% champion JABC and have already brought it forward to other principals.", ""),
        ],
    },
    "RachelTremblayCote_Interview.xlsx": {
        "sheet": "P1 - Teacher-Active",
        "rows": [
            ("C1", "Familiarity & Perception", "How familiar are you with JABC? (1-5)", "5", ""),
            ("C4", "Familiarity & Perception", "Describe JABC to a colleague.",
             "JABC provides ready-made, expert-supported experiential learning tied to student outcomes.", ""),
            ("D2", "Decision-Making", "What makes a program worth classroom time?",
             "Real life connections to career exposure and post-secondary planning.", ""),
            ("E1", "JABC Experience", "Describe your recent JABC experience.",
             "Excellent and credible. The team is enthusiastic and very reliable.", ""),
            ("I3", "Communication & Future Opportunities", "Would you recommend or champion JABC to other educators?",
             "Yes, I would champion it and share it with colleagues whenever I can.", ""),
        ],
    },
    "Cole_Interview.xlsx": {
        "sheet": "P2 - Teacher-Lapsed",
        "rows": [
            ("C1", "Familiarity & Perception", "How familiar are you with JABC? (1-5)", "4", ""),
            ("C4", "Familiarity & Perception", "Describe JABC to a colleague.",
             "It's a solid financial literacy program with hands-on activities for students.", ""),
            ("F1", "Lapsed Engagement", "Have you previously participated in JABC programming?",
             "Yes, we used JABC for a few years and it was a strong program.", ""),
            ("F2", "Lapsed Engagement", "Why did you stop using JABC this year?",
             "We stopped using it because scheduling no longer works for us and registration became difficult "
             "with staff turnover.", ""),
            ("F3", "Lapsed Engagement", "Would you recommend or champion JABC to other educators?",
             "I'd probably recommend it, but I'm not currently championing it since we're not running it.", ""),
        ],
    },
    "AndreaNasiopoulos_Interview.xlsx": {
        "sheet": "P3 - Teacher-Familiar-NoHistory",
        "rows": [
            ("C1", "Familiarity & Perception", "How familiar are you with JABC? (1-5)", "3", ""),
            ("C2", "Familiarity & Perception", "What is the first thing that comes to mind when you hear JABC?",
             "I've heard of JABC but I'm not totally sure what it provides.", ""),
            ("G1", "Never Ran External Program", "Do you use any other external classroom programs?",
             "We currently use a couple of other programs and are open to trying new programming.", ""),
            ("G2", "Never Ran External Program", "What barriers exist to introducing new programs?",
             "Scheduling and workload make it hard to coordinate anything new mid-year.", ""),
            ("I3", "Communication & Future Opportunities", "Would you recommend or champion JABC to other educators?",
             "I would need more information before I could recommend it confidently.", ""),
        ],
    },
    "RoryPayment_Interview.xlsx": {
        "sheet": "P4 - Teacher-Unfamiliar",
        "rows": [
            ("C1", "Familiarity & Perception", "How familiar are you with JABC? (1-5)", "1", ""),
            ("C2", "Familiarity & Perception", "What is the first thing that comes to mind when you hear JABC?",
             "I am not familiar with JABC and don't really know what it does.", ""),
            ("G1", "Never Ran External Program", "Do you use any other external classroom programs?",
             "We've tried a couple of external programs but nothing has really stuck.", ""),
            ("G2", "Never Ran External Program", "What barriers exist to introducing new programs?",
             "Limited capacity at our smaller school and teacher turnover make continuity difficult.", ""),
            ("I3", "Communication & Future Opportunities", "Would you recommend or champion JABC to other educators?",
             "I really can't say, I don't know enough about it to recommend anything yet.", ""),
        ],
    },
    "ErinPaul_Interview.xlsx": {
        "sheet": "P5 - Superintendent-Active",
        "rows": [
            ("C1", "Familiarity & Perception", "How familiar are you with JABC? (1-5)", "5", ""),
            ("C4", "Familiarity & Perception", "Describe JABC to a colleague.",
             "JABC is a robust, curriculum-aligned program delivering real student outcomes across our district.", ""),
            ("H1", "Superintendent / District Lens", "Does JABC align with your district's instructional goals?",
             "Yes, it aligns extremely well with our instructional goals and student achievement priorities.", ""),
            ("E1", "JABC Experience", "Describe your recent JABC experience.",
             "Consistently excellent, the team is professional and highly responsive.", ""),
            ("I3", "Communication & Future Opportunities", "Would you recommend or champion JABC to other educators?",
             "Definitely, I champion JABC across our district and encourage expansion whenever possible.", ""),
        ],
    },
}

QUESTION_BANK_ROWS = [
    ("A1", "Educator Context", "What grade/subject do you teach?", "", ""),
    ("C1", "Familiarity & Perception", "How familiar are you with JABC? (1-5)", "", ""),
]


def build_workbook(path: Path, sheet_name: str, rows: list[tuple]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Question Bank"
    ws.append(HEADERS)
    for row in QUESTION_BANK_ROWS:
        ws.append(list(row))

    ws2 = wb.create_sheet(sheet_name)
    ws2.append(HEADERS)
    for row in rows:
        ws2.append(list(row))

    ref_sheet = wb.create_sheet("Personas & Goals")
    ref_sheet.append(["Persona", "Goals"])
    ref_sheet.append(["Brand Champion", "Deepen engagement"])

    wb.save(path)


def main():
    out_dir = Path(__file__).parent
    for filename, spec in INTERVIEWS.items():
        build_workbook(out_dir / filename, spec["sheet"], spec["rows"])
        print(f"Wrote {filename}")


if __name__ == "__main__":
    main()
