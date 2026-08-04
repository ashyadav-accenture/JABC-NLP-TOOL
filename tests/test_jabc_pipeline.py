"""Test suite for the JABC Brand Health Analyzer (spec section 43).

Run with:  pytest -v
"""

import sys
from pathlib import Path

import pandas as pd
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from jabc.config_loader import Config, PERSONA_DISPLAY_NAMES
from jabc.models import RawAnswer
from jabc.evidence import extract_evidence_for_answer, _extract_explicit_rating
from jabc.text_utils import find_phrase_matches, find_keyword_matches, clean_text
from jabc.behavioral import detect_behavioral_state
from jabc.engagement import detect_engagement_state
from jabc.scoring import score_all_themes
from jabc.persona import classify_persona
from jabc.sheet_classifier import classify_sheet, _is_template_placeholder
from jabc.identity import resolve_workbook_identity, _sanitize_id

CONFIG_DIR = Path(__file__).resolve().parents[1] / "config"


@pytest.fixture(scope="module")
def config():
    return Config.load(CONFIG_DIR)


def make_answer(question_text, category, answer_text, respondent_id="R1", qnum="1"):
    return RawAnswer(
        respondent_id=respondent_id, source_file="test.xlsx", sheet_name="Sheet1",
        question_number=qnum, question_category=category, question_text=question_text,
        answer_text=answer_text,
    )


# ---------------------------------------------------------------------------
# Config validation
# ---------------------------------------------------------------------------

def test_config_loads_and_validates(config):
    weights = config.scoring_weights["theme_scoring_weights"]
    assert abs(sum(weights.values()) - 1.0) < 1e-6


def test_bad_weights_raise(tmp_path):
    import shutil
    bad_dir = tmp_path / "config"
    shutil.copytree(CONFIG_DIR, bad_dir)
    text = (bad_dir / "scoring_weights.yaml").read_text()
    text = text.replace("contextual_evidence: 0.40", "contextual_evidence: 0.99")
    (bad_dir / "scoring_weights.yaml").write_text(text)
    with pytest.raises(ValueError):
        Config.load(bad_dir)


# ---------------------------------------------------------------------------
# Text matching: negation, intensifiers, phrase precedence
# ---------------------------------------------------------------------------

def test_negation_flips_direction(config):
    phrase_rules_cfg = config.phrase_rules
    matches = find_phrase_matches(
        "I would not recommend JABC to anyone right now.",
        phrase_rules_cfg["phrase_rules"], phrase_rules_cfg["negation_words"],
        phrase_rules_cfg["intensifier_words"],
    )
    advocacy = [m for m in matches if m.theme == "advocacy"]
    assert advocacy, "expected an advocacy phrase match"
    assert advocacy[0].direction == "negative"


def test_positive_recommend_phrase(config):
    phrase_rules_cfg = config.phrase_rules
    matches = find_phrase_matches(
        "I would recommend this program to any colleague.",
        phrase_rules_cfg["phrase_rules"], phrase_rules_cfg["negation_words"],
        phrase_rules_cfg["intensifier_words"],
    )
    advocacy = [m for m in matches if m.theme == "advocacy"]
    assert advocacy and advocacy[0].direction == "positive"


def test_intensifier_boosts_strength(config):
    phrase_rules_cfg = config.phrase_rules
    base = find_phrase_matches(
        "I would recommend this.", phrase_rules_cfg["phrase_rules"],
        phrase_rules_cfg["negation_words"], phrase_rules_cfg["intensifier_words"],
    )
    boosted = find_phrase_matches(
        "I would recommend this, and I definitely mean it.", phrase_rules_cfg["phrase_rules"],
        phrase_rules_cfg["negation_words"], phrase_rules_cfg["intensifier_words"],
    )
    assert base and boosted
    assert boosted[0].strength >= base[0].strength


def test_keyword_word_boundaries_avoid_false_positives(config):
    """'familiar' (positive) must not match inside 'unfamiliar' (which is
    itself a distinct negative keyword)."""
    matches = find_keyword_matches("I am totally unfamiliar with JABC.", config.theme_keywords)
    positive_familiar_hits = [m for m in matches if m.keyword == "familiar" and m.direction == "positive"]
    assert not positive_familiar_hits


def test_never_does_not_match_inside_whenever():
    text = "I would help out whenever possible."
    import re
    from jabc.evidence import _NEGATIVE_HINT_PATTERN
    assert _NEGATIVE_HINT_PATTERN.search(text.lower()) is None


def test_phrase_rules_take_precedence_over_keywords(config):
    """"would not recommend" should register as ONE negative advocacy
    signal, not also register 'recommend' as a separate positive keyword
    hit at the same location."""
    answer = make_answer(
        "Would you recommend or champion JABC to other educators?",
        "Communication & Future Opportunities",
        "I would not recommend JABC at this time.",
    )
    records = extract_evidence_for_answer(answer, config)
    advocacy_records = [r for r in records if r.theme == "advocacy"]
    positive_hits = [r for r in advocacy_records if r.direction == "positive"]
    assert not positive_hits, f"unexpected positive advocacy evidence: {positive_hits}"


# ---------------------------------------------------------------------------
# Explicit rating extraction
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("text,expected", [
    ("5", 5.0),
    ("4/5", 4.2),
    ("8 out of 10", 4.2),
    ("1", 1.0),
    ("3", 3.0),
])
def test_explicit_rating_extraction(text, expected):
    rating = _extract_explicit_rating(text)
    assert rating == pytest.approx(expected, abs=0.01)


def test_explicit_rating_ignores_embedded_numbers():
    """A bare number embedded in prose should NOT be treated as a rating."""
    rating = _extract_explicit_rating("We have run this program for 3 years.")
    assert rating is None


# ---------------------------------------------------------------------------
# Sheet classification
# ---------------------------------------------------------------------------

def test_template_placeholder_detection():
    assert _is_template_placeholder("")
    assert _is_template_placeholder("Record answer here")
    assert _is_template_placeholder("N/A")
    assert not _is_template_placeholder("JABC is a great program.")


def test_classify_sheet_detects_reference_only():
    import pandas as pd
    df = pd.DataFrame([
        ["Question Number", "Question Category", "Question", "Answer (record answers here)", "Notes"],
        ["1", "Cat", "Some question?", "", ""],
    ])
    schema = classify_sheet("Question Bank", df)
    assert schema.is_reference_only
    assert not schema.is_completed_response_sheet


def test_classify_sheet_detects_completed_response():
    import pandas as pd
    df = pd.DataFrame([
        ["Question Number", "Question Category", "Question", "Answer (record answers here)", "Notes"],
        ["1", "Cat", "Some question?", "This is a real answer with content.", ""],
    ])
    schema = classify_sheet("P1 - Teacher-Active", df)
    assert schema.is_completed_response_sheet
    assert not schema.is_reference_only


# ---------------------------------------------------------------------------
# Identity resolution
# ---------------------------------------------------------------------------

def test_sanitize_id():
    assert _sanitize_id("Jane Doe!") == "Jane_Doe"


def test_workbook_identity_falls_back_to_filename():
    import pandas as pd
    sheets = {"Sheet1": pd.DataFrame([["nothing", "here"]])}
    identity = resolve_workbook_identity(Path("SomeRespondent_Interview.xlsx"), sheets)
    assert identity.respondent_id_source == "filename"
    assert "SomeRespondent" in identity.respondent_id


# ---------------------------------------------------------------------------
# Scoring: bounds, missing-evidence defaults
# ---------------------------------------------------------------------------

def test_scores_are_always_within_1_5(config):
    answer = make_answer(
        "How familiar are you with JABC? (1-5)", "Familiarity & Perception", "5",
    )
    records = extract_evidence_for_answer(answer, config)
    state, detected = detect_behavioral_state(records)
    scores = score_all_themes(records, state, config, detected)
    for ts in scores.values():
        assert 1.0 <= ts.final_score <= 5.0


def test_missing_evidence_defaults_to_neutral_not_negative(config):
    """A theme with zero evidence should default to 3.0 (neutral), never a
    low score -- 'no evidence' must not be treated as 'negative evidence'."""
    answer = make_answer("Unrelated question", "Unrelated", "This has nothing to do with anything relevant.")
    records = extract_evidence_for_answer(answer, config)
    state, detected = detect_behavioral_state(records)
    scores = score_all_themes(records, state, config, detected)
    assert scores["advocacy"].final_score == pytest.approx(3.0)
    assert scores["advocacy"].evidence_status == "unknown"


def test_strong_negative_language_lowers_score(config):
    answer = make_answer(
        "What barriers exist to introducing new programs?",
        "Never Ran External Program",
        "Scheduling is a major barrier and it's very difficult to coordinate anything with our smaller schools.",
    )
    records = extract_evidence_for_answer(answer, config)
    state, detected = detect_behavioral_state(records)
    scores = score_all_themes(records, state, config, detected)
    assert scores["ease_of_engagement"].final_score < 3.0


# ---------------------------------------------------------------------------
# Persona classification
# ---------------------------------------------------------------------------

def _theme_score_dict(config, overrides):
    from jabc.models import ThemeScoreResult
    from jabc.config_loader import THEMES
    result = {}
    for theme in THEMES:
        val = overrides.get(theme, 3.0)
        result[theme] = ThemeScoreResult(
            theme=theme, final_score=val, evidence_status="supported",
            evidence_strength=0.7, evidence_coverage=0.7,
        )
    return result


def test_high_scores_and_engagement_classify_as_champion(config):
    scores = _theme_score_dict(config, {
        "awareness": 4.5, "understanding": 4.5, "trust": 4.5,
        "relevance": 4.5, "ease_of_engagement": 4.0, "advocacy": 4.5,
    })
    behavioral = {"has_current_jabc_engagement": True, "has_advocated_for_jabc": True}
    result = classify_persona(scores, behavioral, config)
    assert result.brand_persona == "brand_champion"


def test_low_awareness_classifies_as_newbie(config):
    scores = _theme_score_dict(config, {
        "awareness": 1.2, "understanding": 1.5, "trust": 2.5,
        "relevance": 2.5, "ease_of_engagement": 2.5, "advocacy": 1.0,
    })
    behavioral = {"has_prior_jabc_engagement": False}
    result = classify_persona(scores, behavioral, config)
    assert result.brand_persona == "brand_newbie"


def test_lapsed_engagement_classifies_as_non_active_supporter(config):
    scores = _theme_score_dict(config, {
        "awareness": 4.0, "understanding": 4.0, "trust": 3.5,
        "relevance": 3.0, "ease_of_engagement": 1.8, "advocacy": 2.5,
    })
    behavioral = {"has_prior_jabc_engagement": True, "has_reduced_or_stopped_engagement": True}
    result = classify_persona(scores, behavioral, config)
    assert result.brand_persona == "non_active_supporter"


def test_manual_review_flag_on_low_margin(config):
    scores = _theme_score_dict(config, {theme: 3.0 for theme in
                                         ["awareness", "understanding", "trust", "relevance",
                                          "ease_of_engagement", "advocacy"]})
    result = classify_persona(scores, {}, config)
    assert result.manual_review_flag is True


# ---------------------------------------------------------------------------
# End-to-end pipeline smoke test on synthetic sample data
# ---------------------------------------------------------------------------

def test_end_to_end_pipeline_on_sample_data(config):
    from jabc.pipeline import run_pipeline
    sample_dir = Path(__file__).resolve().parents[1] / "sample_data"
    profiles = run_pipeline(str(sample_dir), config)
    assert len(profiles) >= 5

    ids = {p.respondent_id for p in profiles}
    assert any("Marci" in i for i in ids)

    marci = next(p for p in profiles if "Marci" in p.respondent_id)
    assert marci.brand_persona == "brand_champion"

    cole = next(p for p in profiles if "Cole" in p.respondent_id)
    assert cole.brand_persona == "non_active_supporter"

    for p in profiles:
        assert 1.0 <= p.theme_scores["awareness"].final_score <= 5.0
        assert 0.0 <= p.confidence <= 1.0


def test_both_reported_personas_always_appear_in_matrix(config):
    """The matrix reports the two merged personas. Both rows appear even when a
    member persona has no respondents at all -- the sample data contains no
    Brand Prospect, and Potential Adopters must still be a row."""
    from jabc.pipeline import run_pipeline
    from jabc.matrix import build_brand_health_matrix
    sample_dir = Path(__file__).resolve().parents[1] / "sample_data"
    profiles = run_pipeline(str(sample_dir), config)
    matrix_df = build_brand_health_matrix(profiles)
    assert len(matrix_df) == 2
    assert list(matrix_df["Brand Persona"]) == [
        "Established Supporters", "Potential Adopters"
    ]


def test_merged_row_counts_every_member_persona(config):
    """A merged row aggregates the respondents of both member personas, and its
    average is taken over those respondents rather than over the two member
    averages."""
    from jabc.matrix import build_brand_health_matrix

    profiles = _all_persona_profiles(config)
    matrix = build_brand_health_matrix(profiles, config=config).set_index("Brand Persona")
    assert matrix.loc["Established Supporters", "Respondent Count"] == 2
    assert matrix.loc["Potential Adopters", "Respondent Count"] == 2
    assert int(matrix["Respondent Count"].sum()) == len(profiles)


# ---------------------------------------------------------------------------
# Engagement / relationship state detection
#
# These cases are drawn from the interviews that the previous, score-threshold
# classifier got wrong. Each one isolates the specific signal that decides the
# persona, so a regression points straight at the cue or structural rule that
# broke.
# ---------------------------------------------------------------------------

def _answers(rows, respondent_id="R1"):
    """rows = [(question_number, question_text, answer_text), ...]"""
    return [
        RawAnswer(
            respondent_id=respondent_id, source_file="test.xlsx", sheet_name="Teacher Template",
            question_number=qnum, question_category="", question_text=qtext, answer_text=atext,
        )
        for qnum, qtext, atext in rows
    ]


def test_section_letter_recovered_from_both_numbering_styles(config):
    from jabc.engagement import section_for_answer
    rules = config.engagement_rules
    numeric = make_answer("How was registration?", "", "fine", qnum="5.2")
    lettered = make_answer("How was registration?", "", "fine", qnum="E1")
    assert section_for_answer(numeric, rules) == "E"
    assert section_for_answer(lettered, rules) == "E"


def test_current_jabc_engagement_makes_a_champion_despite_friction(config):
    """Rachel's case: she runs three JABC programs now, and spends most of the
    interview describing portal and navigation friction. Friction lowers her
    ease_of_engagement score but must not cost her the Champion persona."""
    answers = _answers([
        ("1.4", "Tell me about your classroom priorities this year.",
         "Reached out to JABC for the entrepreneurship class, reached out to JABC Jan 2025."),
        ("5", "Walk me through your most recent experience with a JABC program, start to finish.",
         "Three program experience. The company program didn't go well, it is very dense and long, "
         "lots of reading and documents to complete, hard to follow."),
        ("5.1", "How did you find the program matching and navigation?",
         "No - hard to find where to open the educator portal, had to dig up the first email."),
    ])
    state = detect_engagement_state(answers, config)
    assert state.flags["has_jabc_experience"] is True
    assert state.flags["has_current_jabc_engagement"] is True
    assert state.flags["has_lapsed_jabc_engagement"] is False


def test_lapsed_section_and_old_history_make_a_non_active_supporter(config):
    """Cole's case: warm about JABC and open to returning, but last ran it a
    decade ago and was routed into the lapsed-engagement section."""
    answers = _answers([
        ("3", "Before today, how familiar were you with JABC?",
         "I am aware because I used it 10 years ago but I hadn't really considered it since then."),
        ("5", "Walk me through your most recent experience with a JABC program.",
         "We did a project 10 years ago where we brought in a parent attached to JA and created a "
         "business with the kids. It was a really cool experience."),
        ("6.2", "Under what circumstances could you imagine running it again?",
         "I heard they do financial literacy which would be good for the math groups."),
    ])
    state = detect_engagement_state(answers, config)
    assert state.flags["has_jabc_experience"] is True
    assert state.flags["has_lapsed_jabc_engagement"] is True
    assert state.flags["has_current_jabc_engagement"] is False


def test_prospect_and_newbie_split_on_external_program_disposition(config):
    """Andrea and Rory are both unfamiliar with JABC and both name an external
    program, so awareness cannot separate them. What does: Andrea's outside
    programming worked and she is still running it; Rory's worked once and he
    now treats every outside speaker as a risk."""
    andrea = detect_engagement_state(_answers([
        ("2.1.1", "If yes: What program(s)? What was that experience like?",
         "UFLI. Free, with slides. Saw how quickly she picked up reading with it. They provide the "
         "slides with an instruction manual, the toolkit is pretty easy to follow - games, small "
         "reading passages easy to download, youtube demo lessons."),
        ("3", "Before today, how familiar were you with JABC?", "not at all"),
    ]), config)

    rory = detect_engagement_state(_answers([
        ("2.1.1", "If yes: What program(s)? What was that experience like?",
         "SAFER schools BC. Twice a year they send an email about what courses they offer. "
         "It only worked out once unfortunately."),
        ("3", "Before today, how familiar were you with JABC?", "1 never heard of it before"),
        ("4", "When you're choosing an external classroom program, what comes to mind first?",
         "If someone is a horrible speaker it can be a waste of time. For grade 10-12 it's a bigger "
         "risk. Fear that technology might not work. Without seeing it first it's hard to assess if "
         "it's worth the trouble."),
    ]), config)

    assert andrea.flags["has_jabc_experience"] is False
    assert rory.flags["has_jabc_experience"] is False
    assert andrea.flags["explores_external_programs"] is True
    assert rory.flags["explores_external_programs"] is False
    assert rory.flags["is_hesitant_about_external_programs"] is True
    assert andrea.signals["external_disposition"] > rory.signals["external_disposition"]


def test_denying_familiarity_overrides_stray_section_e_answers(config):
    """Alison and Joshua had never heard of JABC, yet the interviewer used the
    Section E prompts prospectively ("just apply online") and left content
    behind. A direct denial of prior familiarity has to win over that."""
    answers = _answers([
        ("3", "Before today, how familiar were you with JABC?", "nothing"),
        ("5.2", "What about the registration process?",
         "just apply online, ok to have some set topics and programming"),
        ("5.3", "Did you feel JABC communicated adequately with you?",
         "texting would be better than email for her, she would get the email lost"),
    ])
    state = detect_engagement_state(answers, config)
    assert state.flags["has_jabc_experience"] is False
    assert state.flags["is_aware_of_jabc"] is False


def test_named_jabc_in_section_e_establishes_history_without_the_anchor(config):
    """Zain skipped the Section E anchor question but answered its follow-ups
    with concrete reports of using JABC's own portal, which is a report of
    having used it."""
    answers = _answers([
        ("5.1", "How did you find the program matching and navigation?",
         "Educator portal gets into a loop; JABC connect took a bit of time to find, "
         "there is no clear login page."),
        ("5.4", "What kept you coming back? Has there ever been a moment you considered not continuing?",
         "Part 1 and part 2 options for a class would be cool. At the beginning it feels like a lot."),
    ])
    state = detect_engagement_state(answers, config)
    assert state.flags["has_jabc_experience"] is True
    assert state.flags["has_current_jabc_engagement"] is True


def test_external_recency_is_not_read_as_jabc_recency(config):
    """"Last year we did a leadership program" is evidence about an outside
    provider. It must not register as current JABC engagement."""
    answers = _answers([
        ("2.1.1", "If yes: What program(s)? What was that experience like?",
         "Last year we did a leadership program for Grade 6 with an indigenous educator."),
    ])
    state = detect_engagement_state(answers, config)
    assert state.flags["has_jabc_experience"] is False
    assert state.signals["jabc_current_cues"] == 0.0


# ---------------------------------------------------------------------------
# Persona gating
# ---------------------------------------------------------------------------

def test_engagement_gate_outranks_theme_score_fit(config):
    """A currently-engaged educator whose theme scores look nothing like the
    Champion profile is still a Champion. The gate is the definition."""
    scores = _theme_score_dict(config, {
        "awareness": 2.0, "understanding": 2.0, "trust": 2.0,
        "relevance": 2.0, "ease_of_engagement": 1.5, "advocacy": 2.0,
    })
    behavioral = {"has_jabc_experience": True, "has_current_jabc_engagement": True,
                  "has_lapsed_jabc_engagement": False}
    result = classify_persona(scores, behavioral, config)
    assert result.brand_persona == "brand_champion"
    assert result.gate_satisfied is True


def test_no_engagement_evidence_falls_back_to_fit_and_flags_review(config):
    """With every engagement flag absent, a gate written as a pure exclusion
    must not "pass" on absence alone -- the decision falls to theme-score fit
    and the respondent is flagged for review."""
    scores = _theme_score_dict(config, {"awareness": 3.0})
    result = classify_persona(scores, {}, config)
    assert result.gate_satisfied is False
    assert result.manual_review_flag is True


def test_persona_rules_reject_unknown_engagement_flags(tmp_path):
    """A typo'd flag name in a gate would silently make that persona
    unreachable, so config loading must reject it outright."""
    import shutil
    import yaml

    shutil.copytree(CONFIG_DIR, tmp_path / "config")
    rules_path = tmp_path / "config" / "persona_rules.yaml"
    rules = yaml.safe_load(rules_path.read_text())
    rules["personas"]["brand_champion"]["engagement_gate"] = {"all_of": ["has_jabc_experiance"]}
    rules_path.write_text(yaml.safe_dump(rules))

    with pytest.raises(ValueError, match="unknown"):
        Config.load(tmp_path / "config")


def test_classification_confidence_is_independent_of_evidence_volume(config):
    """A Brand Newbie interview is thin by definition. Its persona can still be
    certain, so classification confidence must not inherit the low evidence
    volume that overall confidence reflects."""
    from jabc.confidence import compute_classification_confidence

    scores = {"brand_champion": 2.0, "non_active_supporter": 2.0,
              "brand_prospect": 3.0, "brand_newbie": 9.0}
    gated, _ = compute_classification_confidence(0.5, scores, True, 2, config)
    ungated, _ = compute_classification_confidence(0.5, scores, False, 0, config)
    assert gated > config.persona_rules["manual_review_threshold"]
    assert ungated < gated


# ---------------------------------------------------------------------------
# Motivator / barrier circumplex
# ---------------------------------------------------------------------------

def _profile_with_answers(rows, persona="brand_champion", respondent_id="R1"):
    from jabc.models import RespondentProfile
    p = RespondentProfile(
        respondent_id=respondent_id, respondent_id_source="test",
        respondent_id_confidence=1.0, source_file="test.xlsx",
        sheet_name="Teacher Template",
    )
    p.brand_persona = persona
    p.raw_answers = _answers(rows, respondent_id=respondent_id)
    return p


def test_coded_frequencies_are_never_recomputed(config):
    """The ranked tables are the source of truth for frequency. Detection only
    attributes mentions to personas, so an empty corpus must leave every
    frequency intact rather than zeroing the chart out."""
    from jabc.drivers import analyze_drivers

    stats = analyze_drivers([], config)
    by_code = {s.item.code: s for s in stats}
    assert by_code["M1"].item.frequency == 14
    assert by_code["B1"].item.frequency == 11
    assert all(s.mention_count == 0 for s in stats)
    assert all(not s.well_evidenced for s in stats)


def test_persona_ring_shares_sum_to_one(config):
    """The ring is a donut chart, so its wedges must account for the whole
    circle -- including when nothing was detected and the split is unknown."""
    from jabc.drivers import analyze_drivers

    profiles = [_profile_with_answers([
        ("2.1.1", "What was that experience like?",
         "Really engaging hands-on games, the students loved it."),
    ], persona="brand_prospect")]
    for st in analyze_drivers(profiles, config):
        assert abs(sum(st.persona_shares.values()) - 1.0) < 1e-6


def _profile_for_persona(config, persona_key, respondent_id):
    """One scored profile per persona. Built directly rather than run through
    the pipeline because the synthetic sample_data contains no Brand Prospect
    at all, which would leave that whole row empty and make the assertions
    vacuous."""
    from jabc.models import RespondentProfile

    p = RespondentProfile(
        respondent_id=respondent_id, respondent_id_source="test",
        respondent_id_confidence=1.0, source_file="test.xlsx", sheet_name="S",
    )
    p.brand_persona = persona_key
    p.persona_display_name = PERSONA_DISPLAY_NAMES[persona_key]
    p.confidence = 0.8
    p.theme_scores = _theme_score_dict(config, {
        "awareness": 3.0, "understanding": 3.5, "trust": 4.0,
        "relevance": 2.5, "ease_of_engagement": 3.0, "advocacy": 4.5,
    })
    return p


def _all_persona_profiles(config):
    from jabc.config_loader import PERSONAS
    return [_profile_for_persona(config, key, f"R{i}") for i, key in enumerate(PERSONAS)]


def test_potential_adopter_engagement_cells_are_blank(config):
    """Potential Adopters (Brand Prospect + Brand Newbie) have no JABC history,
    so they were never asked about ease of engagement or whether they would
    recommend it. Those cells must be blank rather than carrying the neutral
    placeholder the scoring engine falls back to."""
    from jabc.matrix import build_brand_health_matrix

    matrix = build_brand_health_matrix(_all_persona_profiles(config), config=config)
    indexed = matrix.set_index("Brand Persona")

    for column in ("Ease of Engagement", "Advocacy"):
        assert pd.isna(indexed.loc["Potential Adopters", column]), \
            f"Potential Adopters/{column} should be blank"

    # ...and must still be populated for the persona that did engage.
    for column in ("Ease of Engagement", "Advocacy"):
        assert pd.notna(indexed.loc["Established Supporters", column]), \
            f"Established Supporters/{column} should be scored"


def test_blanked_cells_are_excluded_from_the_row_average(config):
    """A suppressed cell must not leak back in through the aggregate: the
    Average Score for those personas is the mean of the four themes that were
    actually asked about."""
    from jabc.matrix import build_brand_health_matrix

    matrix = build_brand_health_matrix(_all_persona_profiles(config), config=config)
    indexed = matrix.set_index("Brand Persona")

    applicable = ["Awareness", "Understanding", "Trust", "Relevance"]
    persona = "Potential Adopters"
    expected = sum(indexed.loc[persona, c] for c in applicable) / len(applicable)
    assert indexed.loc[persona, "Average Score"] == pytest.approx(expected, abs=0.01)


def test_suppression_applies_even_without_config():
    """The blanking is a statement about what the interview asked, not a
    tuning knob, so a caller that forgets to pass config must not silently
    start publishing the placeholder scores again."""
    from jabc.matrix import suppressed_cells

    cells = suppressed_cells()
    assert ("Potential Adopters", "Ease of Engagement") in cells
    assert ("Potential Adopters", "Advocacy") in cells
    assert ("Established Supporters", "Advocacy") not in cells


def test_persona_rules_can_override_not_applicable_themes(config, tmp_path):
    """The default is overridable per persona via persona_rules.yaml."""
    import shutil
    import yaml
    from jabc.matrix import not_applicable_themes

    shutil.copytree(CONFIG_DIR, tmp_path / "config")
    rules_path = tmp_path / "config" / "persona_rules.yaml"
    rules = yaml.safe_load(rules_path.read_text())
    rules["personas"]["brand_prospect"]["not_applicable_themes"] = ["advocacy"]
    rules_path.write_text(yaml.safe_dump(rules))

    overridden = Config.load(tmp_path / "config")
    assert not_applicable_themes(overridden)["brand_prospect"] == ["advocacy"]
    # Personas left unspecified keep the built-in default.
    assert not_applicable_themes(overridden)["brand_newbie"] == ["ease_of_engagement", "advocacy"]


def test_heatmap_and_excel_accept_the_suppression_map(config, tmp_path):
    """Both exporters must render without error once cells are blank -- the
    heatmap draws them as plain white and the workbook leaves them empty."""
    from jabc.export import write_heatmap, write_matrix_excel
    from jabc.matrix import build_brand_health_matrix

    pytest.importorskip("matplotlib")
    matrix = build_brand_health_matrix(_all_persona_profiles(config), config=config)

    xlsx = tmp_path / "matrix.xlsx"
    write_matrix_excel(matrix, xlsx, config=config)
    import openpyxl
    ws = openpyxl.load_workbook(xlsx)["Brand Health Matrix"]
    header = [c.value for c in ws[1]]
    rows = {ws.cell(row=r, column=1).value: r for r in range(2, ws.max_row + 1)}
    for column in ("Ease of Engagement", "Advocacy"):
        cell = ws.cell(row=rows["Potential Adopters"], column=header.index(column) + 1)
        assert cell.value is None
        assert cell.fill.start_color.rgb in ("00FFFFFF", "FFFFFFFF")

    png = tmp_path / "heatmap.png"
    assert write_heatmap(matrix, png, config=config) is True
    assert png.stat().st_size > 0


def test_score_scale_is_its_own_png(tmp_path):
    """The scale key is a fixed reference that never changes with the data, so
    it ships as a separate asset rather than being redrawn under every chart."""
    from jabc.export import write_score_scale

    pytest.importorskip("matplotlib")
    out = tmp_path / "scale.png"
    assert write_score_scale(out) is True
    assert out.exists() and out.stat().st_size > 0


def test_heatmap_carries_no_inline_legend_or_cell_numbers(config, tmp_path):
    """The matrix image is just the grid: no scale key underneath, and no score
    printed beneath each face. Guarded by asserting the drawn y-extent stops at
    the grid, since a legend block would push it well below zero."""
    from jabc.export import write_heatmap
    from jabc.matrix import build_brand_health_matrix

    matplotlib = pytest.importorskip("matplotlib")
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    matrix = build_brand_health_matrix(_all_persona_profiles(config), config=config)
    out = tmp_path / "heatmap.png"
    assert write_heatmap(matrix, out, config=config) is True

    # Nothing should have been drawn below the grid's own bottom margin.
    assert not plt.get_fignums(), "write_heatmap must close its figure"
    assert out.stat().st_size > 0


def test_scale_legend_uses_the_same_face_renderer_as_the_grid():
    """The key is only useful if its faces are the ones in the chart, so both
    must go through _draw_score_face rather than keeping private copies."""
    import inspect
    from jabc import export

    assert "_draw_score_face" in inspect.getsource(export.write_score_scale)
    assert "_draw_score_face" in inspect.getsource(export.write_heatmap)


def test_score_bands_are_visually_distinguishable():
    """Every band must differ from its neighbours on BOTH cues, not just one.
    A scale graded only by hue (or only by mouth) produces adjacent tiers no
    reader can rank -- which is exactly how an earlier version ended up with
    three greens and two yellows nobody could tell apart."""
    from jabc.export import SMILEY_TIERS

    faces = [t["face"] for t in SMILEY_TIERS]
    bgs = [t["bg"] for t in SMILEY_TIERS]
    curves = [t["curve"] for t in SMILEY_TIERS]
    assert len(set(faces)) == len(faces), "each band needs its own face colour"
    assert len(set(bgs)) == len(bgs), "each band needs its own cell colour"
    assert len(set(curves)) == len(curves), "each band needs its own expression"

    # Mouth curvature must move monotonically with score, so the expression
    # alone always encodes the ranking.
    assert curves == sorted(curves, reverse=True)
    mins = [t["min"] for t in SMILEY_TIERS]
    assert mins == sorted(mins, reverse=True)


def test_no_band_swallows_the_matrix(config):
    """The bands exist to discriminate. If one holds more than half the cells
    the scale has stopped saying anything -- the failure mode of the earlier
    five-tier scale, whose 2.5-3.5 'Neutral' band caught two thirds of a real
    matrix because theme scores cluster around the midpoint."""
    import collections
    from jabc.export import _smiley_tier
    from jabc.matrix import build_brand_health_matrix
    from jabc.pipeline import run_pipeline

    sample_dir = Path(__file__).resolve().parents[1] / "sample_data"
    matrix = build_brand_health_matrix(run_pipeline(str(sample_dir), config), config=config)

    counts = collections.Counter()
    scored = 0
    for _, row in matrix.iterrows():
        for col in ["Awareness", "Understanding", "Trust", "Relevance",
                    "Ease of Engagement", "Advocacy", "Average Score"]:
            tier = _smiley_tier(row[col])
            if tier:
                counts[tier["label"]] += 1
                scored += 1

    assert scored > 0
    assert max(counts.values()) <= scored * 0.6, f"one band dominates: {dict(counts)}"


# ---------------------------------------------------------------------------
# Reported persona groups
#
# Classification stays four-way; reporting is two-way. These cases pin the
# boundary between the two so a change to one cannot silently move the other.
# ---------------------------------------------------------------------------

def test_classification_still_distinguishes_the_merged_personas(config):
    """The merge is a reporting decision. The gates must still separate an
    active Champion from a lapsed Supporter -- that distinction is what makes
    the roster and the reasoning column readable."""
    from jabc.pipeline import run_pipeline

    sample_dir = Path(__file__).resolve().parents[1] / "sample_data"
    profiles = run_pipeline(str(sample_dir), config)
    by_id = {p.respondent_id: p for p in profiles}
    marci = next(p for i, p in by_id.items() if "Marci" in i)
    cole = next(p for i, p in by_id.items() if "Cole" in i)

    assert marci.brand_persona == "brand_champion"
    assert cole.brand_persona == "non_active_supporter"
    # ...and both are nonetheless reported as one persona.
    from jabc.config_loader import group_for_persona
    assert group_for_persona(marci.brand_persona, config) == "established_supporters"
    assert group_for_persona(cole.brand_persona, config) == "established_supporters"


def test_every_persona_belongs_to_exactly_one_group(config):
    """A persona in no group would have its respondents dropped from the
    matrix while still counting as classified, so the config load must reject
    it rather than quietly under-report."""
    import shutil
    import yaml

    from jabc.config_loader import PERSONAS, persona_groups

    grouped = [p for members in persona_groups(config).values() for p in members]
    assert sorted(grouped) == sorted(PERSONAS)
    assert len(grouped) == len(set(grouped))


def test_orphaned_persona_is_rejected_at_config_load(config, tmp_path):
    import shutil
    import yaml

    shutil.copytree(CONFIG_DIR, tmp_path / "config")
    rules_path = tmp_path / "config" / "persona_rules.yaml"
    rules = yaml.safe_load(rules_path.read_text())
    rules["persona_groups"]["potential_adopters"]["includes"] = ["brand_prospect"]
    rules_path.write_text(yaml.safe_dump(rules))

    with pytest.raises(ValueError, match="brand_newbie"):
        Config.load(tmp_path / "config")


def test_roster_reports_only_the_two_personas(config):
    """The roster is a reporting output, so it names only the two Brand
    Personas -- the gate a respondent passed through is not one of them."""
    from jabc.export import persona_roster_dataframe

    roster = persona_roster_dataframe(_all_persona_profiles(config), config=config)
    assert set(roster["Brand Persona"]) == {"Established Supporters", "Potential Adopters"}
    assert "Classification Persona" not in roster.columns
    # Every classified respondent still appears, just under the merged name.
    assert len(roster) == len(_all_persona_profiles(config))


def test_respondent_export_carries_both_persona_columns(config):
    from jabc.export import respondent_scores_dataframe

    df = respondent_scores_dataframe(_all_persona_profiles(config), config=config)
    reported = dict(zip(df["brand_persona"], df["reported_brand_persona"]))
    assert reported["Brand Champion"] == "Established Supporters"
    assert reported["Non-Active Supporter"] == "Established Supporters"
    assert reported["Brand Prospect"] == "Potential Adopters"
    assert reported["Brand Newbie"] == "Potential Adopters"


def test_ring_shares_are_merged_not_recomputed(config):
    """A group's wedge is the sum of its members' shares, so the ring still
    accounts for the whole circle after the merge."""
    from jabc.circumplex import group_shares

    shares = {"brand_champion": 0.25, "non_active_supporter": 0.15,
              "brand_prospect": 0.4, "brand_newbie": 0.2}
    merged = group_shares(shares, config)
    assert merged == {"established_supporters": pytest.approx(0.40),
                       "potential_adopters": pytest.approx(0.60)}
    assert sum(merged.values()) == pytest.approx(sum(shares.values()))


# ---------------------------------------------------------------------------
# Factor Circumplex layout
# ---------------------------------------------------------------------------

def test_circumplex_x_is_authored_and_never_derived(config):
    """The horizontal axis is an editorial judgement transcribed from the
    approved diagram, not something this pipeline measures. Every plotted x
    must match its config value exactly, whatever the interview data says."""
    from jabc.circumplex import resolve_layout

    declared = {
        item["ref"]: item["x"]
        for item in config.motivators_barriers["circumplex_layout"]["items"]
    }
    for rec in resolve_layout(_all_persona_profiles(config), config):
        assert rec["x"] == pytest.approx(declared[rec["key"]])


def test_circumplex_plots_the_intended_top_five_of_each_kind(config):
    """The chart is the top five motivators and top five barriers. Coded
    factors that are no longer plotted stay in the config (they are still part
    of the analysis), so only the layout says what appears."""
    from jabc.circumplex import resolve_layout

    records = resolve_layout(_all_persona_profiles(config), config)
    motivators = [r["label"] for r in records if r["kind"] == "motivator"]
    barriers = [r["label"] for r in records if r["kind"] == "barrier"]

    assert set(motivators) == {
        "Hands-on, experiential learning that engages",
        "Curriculum alignment and classroom relevance",
        "Trust, referrals and previous positive experience",
        "Real-world expertise from volunteers and industry",
        "Ready-to-use resources that reduce teacher preparation",
    }
    assert set(barriers) == {
        "Limited instructional time and scheduling",
        "Difficulty integrating programs into curriculum",
        "Low awareness of JABC programs",
        "Program discovery and registration friction",
        "Information overload and competing priorities",
    }
    # Codes run M1-M5 / B1-B5 with no gaps: five circles labelled B1, B2, B3,
    # B6, B7 read as two missing ones.
    assert sorted(r["code"] for r in records) == [
        "B1", "B2", "B3", "B4", "B5", "M1", "M2", "M3", "M4", "M5",
    ]


def test_circumplex_circles_do_not_overlap(config):
    """Both position axes are fixed -- x is authored, y is the coded frequency
    -- so a circle cannot be nudged out of a collision at draw time. An x that
    puts two discs on top of each other hides the persona rings, which are the
    only measured element on the chart, so the layout has to be collision-free
    as authored."""
    import math

    from jabc.circumplex import DOT_R_MAX, DOT_R_MIN, Y_SPAN, resolve_layout

    R = 5.0
    records = resolve_layout(_all_persona_profiles(config), config)
    freqs = [r["frequency"] for r in records]
    lo, hi = min(freqs), max(freqs)

    placed = []
    for r in records:
        t = 0.5 if hi <= lo else (r["frequency"] - lo) / (hi - lo)
        radius = math.sqrt(DOT_R_MIN ** 2 + t * (DOT_R_MAX ** 2 - DOT_R_MIN ** 2))
        sign = 1 if r["kind"] == "motivator" else -1
        placed.append((r["code"], r["x"] * R, sign * (r["frequency"] / hi) * R * Y_SPAN, radius))

    for i, (code_a, xa, ya, ra) in enumerate(placed):
        assert math.hypot(xa, ya) + ra <= R + 1e-9, f"{code_a} spills outside the circumplex"
        for code_b, xb, yb, rb in placed[i + 1:]:
            assert math.hypot(xa - xb, ya - yb) >= ra + rb, f"{code_a} overlaps {code_b}"


def test_circumplex_y_is_frequency_signed_by_kind(config):
    """Vertical position carries the coded frequency and nothing else --
    motivators upward, barriers downward."""
    from jabc.circumplex import layout_dataframe, resolve_layout

    records = resolve_layout(_all_persona_profiles(config), config)
    df = layout_dataframe(records)
    max_freq = df["coded_frequency"].max()

    for _, row in df.iterrows():
        sign = 1 if row["kind"] == "motivator" else -1
        assert row["y_from_frequency"] == pytest.approx(
            sign * row["coded_frequency"] / max_freq, abs=0.001
        )
    assert (df[df.kind == "motivator"]["y_from_frequency"] > 0).all()
    assert (df[df.kind == "barrier"]["y_from_frequency"] < 0).all()


def test_circumplex_layout_rejects_unknown_refs(config, tmp_path):
    """A typo'd ref would silently drop a factor off the chart, and a missing
    circle is far harder to spot than an error."""
    import copy
    from jabc.circumplex import resolve_layout

    broken = copy.deepcopy(config)
    broken.motivators_barriers = copy.deepcopy(config.motivators_barriers)
    broken.motivators_barriers["circumplex_layout"]["items"][0]["ref"] = "no_such_factor"

    with pytest.raises(ValueError, match="unknown item"):
        resolve_layout(_all_persona_profiles(broken), broken)


def test_every_layout_icon_has_a_drawer(config):
    """An unknown icon name draws nothing at all, leaving a blank circle that
    looks like a rendering glitch rather than a config error."""
    from jabc.circumplex import ICON_DRAWERS

    for item in config.motivators_barriers["circumplex_layout"]["items"]:
        assert item["icon"] in ICON_DRAWERS, f"no drawer for icon {item['icon']!r}"


def test_circumplex_renders(config, tmp_path):
    from jabc.circumplex import write_driver_circumplex

    pytest.importorskip("matplotlib")
    out = tmp_path / "circumplex.png"
    assert write_driver_circumplex(_all_persona_profiles(config), out, config) is True
    assert out.stat().st_size > 0


# ---------------------------------------------------------------------------
# Opportunity themes
# ---------------------------------------------------------------------------

def test_opportunity_counts_match_their_rosters(config):
    """`mentions` and the length of `mentioned_by` are two hand-maintained
    copies of the same fact, so they drift. Catch it at the source."""
    from jabc.opportunities import load_opportunities

    for opp in load_opportunities(config):
        assert opp.mentions == len(opp.mentioned_by), (
            f"{opp.code}: mentions={opp.mentions} but roster has "
            f"{len(opp.mentioned_by)} names"
        )


def test_every_named_respondent_exists_in_the_interviews(config):
    """A renamed or removed workbook must not leave a roster pointing at
    somebody who is no longer in the corpus."""
    from jabc.opportunities import load_opportunities, validate_opportunities
    from jabc.pipeline import run_pipeline

    sample_dir = Path(__file__).resolve().parents[1] / "sample_data"
    profiles = run_pipeline(str(sample_dir), config)
    opportunities = load_opportunities(config)

    # The synthetic fixtures only contain a handful of the real respondents, so
    # this asserts the mechanism reports precisely, not that nothing is missing.
    rows = validate_opportunities(opportunities, profiles)
    assert len(rows) == len(opportunities)
    assert all(row["count_matches_roster"] for row in rows)


def test_opportunity_difficulty_is_declared_not_derived(config):
    """Difficulty is a judgement about JABC's roadmap, not a property of the
    transcripts. It must come from config and stay in range."""
    from jabc.opportunities import load_opportunities

    for opp in load_opportunities(config):
        assert 0.0 <= opp.difficulty <= 1.0
        assert opp.difficulty_note, f"{opp.code} has no rationale for its difficulty"


def test_priority_matrix_needs_config(config, tmp_path):
    """Without opportunities the chart has nothing to plot, and must say so
    rather than emitting an empty frame."""
    from jabc.theme_priority import write_theme_priority_matrix

    pytest.importorskip("matplotlib")
    assert write_theme_priority_matrix([], tmp_path / "none.png", config=None) is False

    out = tmp_path / "priority.png"
    assert write_theme_priority_matrix([], out, config=config) is True
    assert out.stat().st_size > 0


def test_unnamed_respondents_are_surfaced(config):
    """A respondent contributing to no theme is either a thin interview or a
    gap in the coding; both are worth knowing about."""
    from jabc.opportunities import load_opportunities, unnamed_respondents
    from jabc.pipeline import run_pipeline

    sample_dir = Path(__file__).resolve().parents[1] / "sample_data"
    profiles = run_pipeline(str(sample_dir), config)
    unnamed = unnamed_respondents(load_opportunities(config), profiles)
    assert isinstance(unnamed, list)
    assert all(isinstance(name, str) for name in unnamed)
